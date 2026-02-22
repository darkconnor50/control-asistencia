import os
import json
import io
from functools import wraps
from flask import Flask, request, jsonify, send_file, session, redirect
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

ID_HOJA = '159p8vlPVs0Nh1yMjuXdFhUGclXjT6e9BalQ2H8No6dA'
ZONA_HORARIA = timezone(timedelta(hours=-6))  # México Centro

app.secret_key = os.environ.get("SECRET_KEY", "clave_secreta_123")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")

creds_json = os.environ.get("GOOGLE_CREDENTIALS", "").strip().lstrip('\ufeff')
if creds_json:
    creds = Credentials.from_service_account_info(json.loads(creds_json), scopes=SCOPES)
else:
    creds = Credentials.from_service_account_file('credenciales.json', scopes=SCOPES)

cliente = gspread.authorize(creds)
hoja = cliente.open_by_key(ID_HOJA)

hoja_empleados = hoja.worksheet('EMPLEADOS')
hoja_sucursales = hoja.worksheet('SUCURSALES')
hoja_registros = hoja.worksheet('REGISTROS')
hoja_turnos = hoja.worksheet('TURNOS')

def login_requerido(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('autenticado'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = ''
    if request.method == 'POST':
        password = request.form.get('password')
        if password == ADMIN_PASSWORD:
            session['autenticado'] = True
            return redirect('/admin')
        else:
            error = 'Contraseña incorrecta'

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Admin - Login</title>
        <style>
            body {{ font-family: Arial; text-align: center; padding: 60px; background: #f0f0f0; }}
            h1 {{ color: #333; }}
            input {{ width: 300px; padding: 12px; margin: 10px; font-size: 16px; border-radius: 8px; border: 1px solid #ccc; }}
            button {{ width: 300px; padding: 12px; background: #333; color: white; border: none; border-radius: 8px; font-size: 16px; cursor: pointer; }}
            .error {{ color: red; margin-top: 10px; }}
        </style>
    </head>
    <body>
        <h1>Panel de Administración</h1>
        <p>Ingresa la contraseña para continuar</p>
        <form method="POST">
            <input type="password" name="password" placeholder="Contraseña" /><br>
            <button type="submit">ENTRAR</button>
        </form>
        <p class="error">{error}</p>
    </body>
    </html>
    """

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

@app.route('/sucursal/<id_sucursal>')
def pagina_sucursal(id_sucursal):
    sucursales = hoja_sucursales.get_all_records()
    sucursal = next((s for s in sucursales if str(s['ID_SUCURSAL']) == id_sucursal), None)

    if not sucursal:
        return "Sucursal no encontrada", 404

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Control de Asistencia</title>
        <style>
            body {{ font-family: Arial; text-align: center; padding: 30px; background: #f0f0f0; }}
            h1 {{ color: #333; }}
            h2 {{ color: #666; }}
            input {{ width: 80%; padding: 12px; margin: 10px; font-size: 16px; border-radius: 8px; border: 1px solid #ccc; }}
            button {{ width: 80%; padding: 15px; margin: 10px; font-size: 18px; border: none; border-radius: 8px; cursor: pointer; color: white; }}
            .entrada {{ background: #28a745; }}
            .salida {{ background: #dc3545; }}
        </style>
    </head>
    <body>
        <h1>Control de Asistencia</h1>
        <h2>{sucursal['NOMBRE_SUCURSAL']}</h2>
        <div id="vista-nombre" style="display:none">
            <p>Primera vez aquí. Escribe tu nombre:</p>
            <input type="text" id="nombre" placeholder="Tu nombre completo" />
            <button class="entrada" onclick="guardarNombre()">GUARDAR</button>
        </div>
        <div id="vista-principal" style="display:none">
            <p id="saludo"></p>
            <button class="entrada" onclick="registrar('ENTRADA')">ENTRADA</button>
            <button class="salida" onclick="registrar('SALIDA')">SALIDA</button>
        </div>
        <p id="mensaje"></p>
        <script>
            function generarUUID() {{
                return 'xxxxxxxx-xxxx-4xxx-yxxx-xxxxxxxxxxxx'.replace(/[xy]/g, function(c) {{
                    var r = Math.random() * 16 | 0, v = c == 'x' ? r : (r & 0x3 | 0x8);
                    return v.toString(16);
                }});
            }}

            let uuid = localStorage.getItem('uuid_dispositivo');
            if (!uuid) {{
                uuid = generarUUID();
                localStorage.setItem('uuid_dispositivo', uuid);
            }}

            let nombreGuardado = localStorage.getItem('nombre_empleado');
            if (!nombreGuardado) {{
                document.getElementById('vista-nombre').style.display = 'block';
            }} else {{
                document.getElementById('saludo').innerText = 'Hola, ' + nombreGuardado;
                document.getElementById('vista-principal').style.display = 'block';
            }}

            function guardarNombre() {{
                const nombre = document.getElementById('nombre').value;
                if (!nombre) {{ alert('Escribe tu nombre'); return; }}
                localStorage.setItem('nombre_empleado', nombre);
                document.getElementById('vista-nombre').style.display = 'none';
                document.getElementById('saludo').innerText = 'Hola, ' + nombre;
                document.getElementById('vista-principal').style.display = 'block';
            }}

            async function registrar(tipo) {{
                const nombre = localStorage.getItem('nombre_empleado');
                const res = await fetch('/registrar', {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{uuid, nombre, tipo, id_sucursal: '{id_sucursal}'}})
                }});
                const data = await res.json();
                document.getElementById('mensaje').innerText = data.mensaje;
            }}
        </script>
    </body>
    </html>
    """
    return html

@app.route('/registrar', methods=['POST'])
def registrar():
    datos = request.json
    uuid = datos['uuid']

    ip_cliente = request.remote_addr
    sucursales = hoja_sucursales.get_all_records()
    sucursal_valida = next((s for s in sucursales if str(s['ID_SUCURSAL']) == datos['id_sucursal']), None)

    if sucursal_valida:
        prefijo_red = '.'.join(ip_cliente.split('.')[:2])
        ip_servidor = request.host.split(':')[0]
        prefijo_servidor = '.'.join(ip_servidor.split('.')[:2])

        if prefijo_red != prefijo_servidor:
            return jsonify({'mensaje': 'Acceso denegado. No estás en la red de la sucursal.'})

    nombre_input = datos.get('nombre', '')
    tipo = datos['tipo']
    id_sucursal = datos['id_sucursal']

    empleados = hoja_empleados.get_all_records()
    empleado = next((e for e in empleados if e['ID_DISPOSITIVO'] == uuid), None)

    if not empleado:
        if not nombre_input:
            return jsonify({'mensaje': 'Primera vez: ingresa tu nombre completo'})
        hoja_empleados.append_row([uuid, nombre_input, datetime.now(ZONA_HORARIA).strftime('%Y-%m-%d %H:%M:%S'), 1])
        nombre = nombre_input
        id_turno = 1
    else:
        nombre = empleado['NOMBRE']
        id_turno = empleado.get('ID_TURNO', 1)

    sucursal = next((s for s in sucursales if str(s['ID_SUCURSAL']) == id_sucursal), None)

    turnos = hoja_turnos.get_all_records()
    turno = next((t for t in turnos if str(t['ID_TURNO']) == str(id_turno)), None)

    ahora = datetime.now(ZONA_HORARIA)
    retardo = 'NO'
    mensaje_retardo = ''

    if tipo == 'ENTRADA' and turno:
        hora_limite = datetime.strptime(turno['HORA_ENTRADA'], '%H:%M').replace(
            year=ahora.year, month=ahora.month, day=ahora.day,
            tzinfo=ZONA_HORARIA)
        if ahora > hora_limite:
            minutos_tarde = int((ahora - hora_limite).total_seconds() / 60)
            retardo = f'{minutos_tarde} min'
            mensaje_retardo = f' ⚠️ Retardo de {minutos_tarde} minutos'

    hoja_registros.append_row([
        nombre,
        sucursal['NOMBRE_SUCURSAL'],
        ahora.strftime('%Y-%m-%d'),
        ahora.strftime('%H:%M:%S'),
        tipo,
        retardo
    ])

    return jsonify({'mensaje': f'{tipo} registrada para {nombre} a las {ahora.strftime("%H:%M:%S")}{mensaje_retardo}'})

@app.route('/admin')
@login_requerido
def admin():
    empleados = hoja_empleados.get_all_records()
    turnos = hoja_turnos.get_all_records()

    filas_empleados = ""
    for e in empleados:
        opciones_turnos = ""
        for t in turnos:
            selected = 'selected' if str(t['ID_TURNO']) == str(e.get('ID_TURNO', 1)) else ''
            opciones_turnos += f"<option value='{t['ID_TURNO']}' {selected}>{t['NOMBRE_TURNO']}</option>"
        filas_empleados += f"""
        <tr>
            <td>{e['NOMBRE']}</td>
            <td>
                <select onchange="cambiarTurno('{e['ID_DISPOSITIVO']}', this.value)">
                    {opciones_turnos}
                </select>
            </td>
        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Panel Admin</title>
        <style>
            body {{ font-family: Arial; padding: 30px; background: #f0f0f0; }}
            h1 {{ color: #333; text-align: center; }}
            table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; }}
            th {{ background: #333; color: white; padding: 12px; text-align: left; }}
            td {{ padding: 12px; border-bottom: 1px solid #eee; }}
            select {{ padding: 8px; border-radius: 6px; border: 1px solid #ccc; font-size: 14px; }}
            #mensaje {{ text-align: center; margin-top: 20px; font-weight: bold; color: #28a745; }}
            .btn-reporte {{ padding: 12px 24px; background: #007bff; color: white; border: none; border-radius: 8px; font-size: 16px; margin: 10px; cursor: pointer; }}
            .btn-logout {{ display: inline-block; padding: 12px 24px; background: #dc3545; color: white; text-align: center; border-radius: 8px; text-decoration: none; font-size: 16px; margin: 10px; }}
            .botones {{ text-align: center; margin-bottom: 20px; }}
        </style>
    </head>
    <body>
        <h1>Panel de Administración</h1>
        <div class="botones">
            <button onclick="descargarReporte()" class="btn-reporte">📥 Descargar Reporte</button>
            <a href="/logout" class="btn-logout">🚪 Cerrar Sesión</a>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Empleado</th>
                    <th>Turno</th>
                </tr>
            </thead>
            <tbody>
                {filas_empleados}
            </tbody>
        </table>
        <p id="mensaje"></p>
        <script>
            async function descargarReporte() {{
                const res = await fetch('/reporte');
                if (res.ok) {{
                    const blob = await res.blob();
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = 'reporte.xlsx';
                    a.click();
                    window.URL.revokeObjectURL(url);
                }} else {{
                    alert('No hay registros para generar reporte.');
                }}
            }}

            async function cambiarTurno(uuid, idTurno) {{
                const res = await fetch('/cambiar_turno', {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{uuid, id_turno: idTurno}})
                }});
                const data = await res.json();
                document.getElementById('mensaje').innerText = data.mensaje;
            }}
        </script>
    </body>
    </html>
    """
    return html

@app.route('/cambiar_turno', methods=['POST'])
@login_requerido
def cambiar_turno():
    datos = request.json
    uuid = datos['uuid']
    id_turno = datos['id_turno']

    empleados = hoja_empleados.get_all_records()
    for i, e in enumerate(empleados, 2):
        if e['ID_DISPOSITIVO'] == uuid:
            hoja_empleados.update_cell(i, 4, id_turno)
            return jsonify({'mensaje': f'Turno actualizado correctamente para {e["NOMBRE"]}'})

    return jsonify({'mensaje': 'Empleado no encontrado'})

@app.route('/reporte')
@login_requerido
def descargar_reporte():
    registros = hoja_registros.get_all_records()
    if not registros:
        return "No hay registros para generar reporte.", 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    encabezados = ['EMPLEADO', 'SUCURSAL', 'FECHA', 'ENTRADA', 'SALIDA', 'HORAS TRABAJADAS', 'RETARDO']
    for col, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=enc)
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        celda.alignment = Alignment(horizontal='center')

    resumen = {}
    for r in registros:
        clave = f"{r['NOMBRE']}_{r['FECHA']}"
        if clave not in resumen:
            resumen[clave] = {
                'nombre': r['NOMBRE'],
                'sucursal': r['SUCURSAL'],
                'fecha': r['FECHA'],
                'entrada': '-',
                'salida': '-',
                'horas': '-',
                'retardo': '-'
            }
        if r['TIPO'] == 'ENTRADA':
            resumen[clave]['entrada'] = r['HORA']
            resumen[clave]['retardo'] = r.get('RETARDO', 'NO')
        elif r['TIPO'] == 'SALIDA':
            resumen[clave]['salida'] = r['HORA']

    for datos in resumen.values():
        if datos['entrada'] != '-' and datos['salida'] != '-':
            entrada = datetime.strptime(datos['entrada'], '%H:%M:%S')
            salida = datetime.strptime(datos['salida'], '%H:%M:%S')
            diff = salida - entrada
            horas = int(diff.total_seconds() // 3600)
            minutos = int((diff.total_seconds() % 3600) // 60)
            datos['horas'] = f'{horas}h {minutos}m'

    for fila, datos in enumerate(resumen.values(), 2):
        ws.cell(row=fila, column=1, value=datos['nombre'])
        ws.cell(row=fila, column=2, value=datos['sucursal'])
        ws.cell(row=fila, column=3, value=datos['fecha'])
        ws.cell(row=fila, column=4, value=datos['entrada'])
        ws.cell(row=fila, column=5, value=datos['salida'])
        ws.cell(row=fila, column=6, value=datos['horas'])
        celda_retardo = ws.cell(row=fila, column=7, value=datos['retardo'])
        if datos['retardo'] != 'NO' and datos['retardo'] != '-':
            celda_retardo.font = Font(color='FF0000', bold=True)

    for col in ws.columns:
        max_ancho = max(len(str(celda.value or '')) for celda in col)
        ws.column_dimensions[col[0].column_letter].width = max_ancho + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_reporte = datetime.now(ZONA_HORARIA).strftime('%Y-%m-%d_%H-%M')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'reporte_{fecha_reporte}.xlsx'
    )

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)