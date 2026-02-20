import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Configuración
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

ID_HOJA = '159p8vlPVs0Nh1yMjuXdFhUGclXjT6e9BalQ2H8No6dA'

creds = Credentials.from_service_account_file('credenciales.json', scopes=SCOPES)
cliente = gspread.authorize(creds)
hoja = cliente.open_by_key(ID_HOJA)
hoja_registros = hoja.worksheet('REGISTROS')

def generar_reporte():
    registros = hoja_registros.get_all_records()
    
    if not registros:
        print("No hay registros para generar reporte.")
        return

    # Crear carpeta de reportes
    if not os.path.exists('reportes'):
        os.makedirs('reportes')

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # Encabezados
    encabezados = ['EMPLEADO', 'SUCURSAL', 'FECHA', 'ENTRADA', 'SALIDA', 'HORAS TRABAJADAS', 'RETARDO']
    for col, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=enc)
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        celda.alignment = Alignment(horizontal='center')

    # Agrupar registros por empleado y fecha
    resumen = {}
    for r in registros:
        clave = f"{r['NOMBRE']}_{r['FECHA']}"
        if clave not in resumen:
            resumen[clave] = {
                'nombre': r['NOMBRE'],
                'sucursal': r['SUCURSAL'],
                'fecha': r['FECHA'],
                'entrada': '-',
                'salida': '-',
                'horas': '-',
                'retardo': '-'
            }
        if r['TIPO'] == 'ENTRADA':
            resumen[clave]['entrada'] = r['HORA']
            resumen[clave]['retardo'] = r.get('RETARDO', 'NO')
        elif r['TIPO'] == 'SALIDA':
            resumen[clave]['salida'] = r['HORA']

    # Calcular horas trabajadas
    for datos in resumen.values():
        if datos['entrada'] != '-' and datos['salida'] != '-':
            entrada = datetime.strptime(datos['entrada'], '%H:%M:%S')
            salida = datetime.strptime(datos['salida'], '%H:%M:%S')
            diff = salida - entrada
            horas = int(diff.total_seconds() // 3600)
            minutos = int((diff.total_seconds() % 3600) // 60)
            datos['horas'] = f'{horas}h {minutos}m'

    # Escribir datos
    for fila, datos in enumerate(resumen.values(), 2):
        ws.cell(row=fila, column=1, value=datos['nombre'])
        ws.cell(row=fila, column=2, value=datos['sucursal'])
        ws.cell(row=fila, column=3, value=datos['fecha'])
        ws.cell(row=fila, column=4, value=datos['entrada'])
        ws.cell(row=fila, column=5, value=datos['salida'])
        ws.cell(row=fila, column=6, value=datos['horas'])
        celda_retardo = ws.cell(row=fila, column=7, value=datos['retardo'])
        if datos['retardo'] != 'NO' and datos['retardo'] != '-':
            celda_retardo.font = Font(color='FF0000', bold=True)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_ancho = max(len(str(celda.value or '')) for celda in col)
        ws.column_dimensions[col[0].column_letter].width = max_ancho + 4

    # Guardar archivo
    fecha_reporte = datetime.now().strftime('%Y-%m-%d_%H-%M')
    archivo = f'reportes/reporte_quincena_{fecha_reporte}.xlsx'
    wb.save(archivo)
    print(f"Reporte generado: {archivo}")

    # Limpiar registros en Sheets
    hoja_registros.clear()
    hoja_registros.append_row(['NOMBRE', 'SUCURSAL', 'FECHA', 'HORA', 'TIPO', 'RETARDO'])
    print("Registros limpiados en Google Sheets.")

if __name__ == '__main__':
    generar_reporte()
